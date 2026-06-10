import os
import csv
import datetime
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QComboBox, QDateEdit, QPushButton, QFrame, 
                             QMessageBox, QCheckBox, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont

# Excel and PDF Generation Packages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from models.visitor import VisitorModel
from models.employee import EmployeeModel
from models.audit_logs import AuditLogModel

REPORTS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "reports"))

class ReportsTab(QWidget):
    def __init__(self, user_session, parent=None):
        super().__init__(parent)
        self.user_session = user_session
        self.init_ui()

    def init_ui(self):
        # Main Layout
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(20, 20, 20, 20)
        self.main_layout.setSpacing(15)

        # Tab Header Title
        title_label = QLabel("Report Management & Exports")
        title_label.setObjectName("TabTitle")
        self.main_layout.addWidget(title_label)

        # Main Configurations Frame
        config_card = QFrame()
        config_card.setObjectName("MetricCard")
        config_layout = QVBoxLayout(config_card)
        config_layout.setContentsMargins(20, 20, 20, 20)
        config_layout.setSpacing(15)

        # 1. Report Type / Scope
        scope_layout = QHBoxLayout()
        scope_lbl = QLabel("Report Scope *:")
        scope_lbl.setObjectName("FormLabel")
        scope_lbl.setFixedWidth(140)
        
        self.scope_combo = QComboBox()
        self.scope_combo.addItems(["Daily Report (Today)", "Weekly Report (Last 7 Days)", "Monthly Report (Last 30 Days)", "Custom Date Range"])
        self.scope_combo.currentIndexChanged.connect(self.on_scope_changed)
        scope_layout.addWidget(scope_lbl)
        scope_layout.addWidget(self.scope_combo)
        config_layout.addLayout(scope_layout)

        # 2. Date Pickers (visible only when Custom is selected)
        self.date_picker_frame = QFrame()
        date_picker_layout = QHBoxLayout(self.date_picker_frame)
        date_picker_layout.setContentsMargins(0, 0, 0, 0)
        date_picker_layout.setSpacing(10)

        from_lbl = QLabel("Date From:")
        from_lbl.setObjectName("FormLabel")
        from_lbl.setFixedWidth(140)
        date_picker_layout.addWidget(from_lbl)

        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        date_picker_layout.addWidget(self.date_from)

        to_lbl = QLabel("Date To:")
        to_lbl.setObjectName("FormLabel")
        date_picker_layout.addWidget(to_lbl)

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        date_picker_layout.addWidget(self.date_to)

        self.date_picker_frame.setVisible(False) # Default hidden since scope is Daily
        config_layout.addWidget(self.date_picker_frame)

        # 3. Status Filters to Include
        status_layout = QHBoxLayout()
        status_title_lbl = QLabel("Include Statuses:")
        status_title_lbl.setObjectName("FormLabel")
        status_title_lbl.setFixedWidth(140)
        status_layout.addWidget(status_title_lbl)

        self.chk_registered = QCheckBox("Registered")
        self.chk_registered.setChecked(True)
        status_layout.addWidget(self.chk_registered)

        self.chk_checkin = QCheckBox("Checked In")
        self.chk_checkin.setChecked(True)
        status_layout.addWidget(self.chk_checkin)

        self.chk_checkout = QCheckBox("Checked Out")
        self.chk_checkout.setChecked(True)
        status_layout.addWidget(self.chk_checkout)
        
        status_layout.addStretch()
        config_layout.addLayout(status_layout)

        # 4. Action Export Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)

        self.pdf_btn = QPushButton("EXPORT TO PDF")
        self.pdf_btn.setObjectName("PrimaryBtn")
        self.pdf_btn.setFixedHeight(40)
        self.pdf_btn.setCursor(Qt.PointingHandCursor)
        self.pdf_btn.clicked.connect(lambda: self.trigger_export("PDF"))
        btn_layout.addWidget(self.pdf_btn)

        self.excel_btn = QPushButton("EXPORT TO EXCEL")
        self.excel_btn.setObjectName("PrimaryBtn")
        self.excel_btn.setFixedHeight(40)
        self.excel_btn.setCursor(Qt.PointingHandCursor)
        self.excel_btn.clicked.connect(lambda: self.trigger_export("EXCEL"))
        btn_layout.addWidget(self.excel_btn)

        self.csv_btn = QPushButton("EXPORT TO CSV")
        self.csv_btn.setObjectName("SecondaryBtn")
        self.csv_btn.setFixedHeight(40)
        self.csv_btn.setCursor(Qt.PointingHandCursor)
        self.csv_btn.clicked.connect(lambda: self.trigger_export("CSV"))
        btn_layout.addWidget(self.csv_btn)

        config_layout.addLayout(btn_layout)
        self.main_layout.addWidget(config_card)

        # Open Reports Directory helper card
        open_dir_card = QFrame()
        open_dir_card.setObjectName("MetricCard")
        dir_layout = QHBoxLayout(open_dir_card)
        dir_lbl = QLabel("Need to check past exports?")
        dir_lbl.setFont(QFont("Segoe UI", 10))
        dir_layout.addWidget(dir_lbl)

        open_dir_btn = QPushButton("Open Reports Directory")
        open_dir_btn.setObjectName("SecondaryBtn")
        open_dir_btn.setFixedWidth(180)
        open_dir_btn.clicked.connect(self.open_reports_dir)
        dir_layout.addWidget(open_dir_btn)

        self.main_layout.addWidget(open_dir_card)
        self.main_layout.addStretch()

    def on_scope_changed(self, index):
        """Show date pickers only when custom range is chosen."""
        self.date_picker_frame.setVisible(index == 3)

    def open_reports_dir(self):
        """Open the local system reports folder in Windows Explorer."""
        os.makedirs(REPORTS_DIR, exist_ok=True)
        os.startfile(REPORTS_DIR)

    def get_query_filters(self):
        """Generate date bounds and status lists according to configurations."""
        filters = {}
        scope_idx = self.scope_combo.currentIndex()
        today = datetime.date.today()
        
        if scope_idx == 0:  # Daily
            filters['date_from'] = today.strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        elif scope_idx == 1:  # Weekly
            filters['date_from'] = (today - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        elif scope_idx == 2:  # Monthly
            filters['date_from'] = (today - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        else:  # Custom Range
            filters['date_from'] = self.date_from.date().toString("yyyy-MM-dd")
            filters['date_to'] = self.date_to.date().toString("yyyy-MM-dd")
            
        return filters

    def fetch_records_for_report(self):
        """Retrieve and filter records based on status checkboxes and date bounds."""
        filters = self.get_query_filters()
        
        # Pull all matching rows
        # Limit set to 5000 to cover comprehensive exports
        all_logs = VisitorModel.get_visitors_paginated(limit=5000, offset=0, filters=filters)
        
        # Filter in python based on status checkboxes (Registered, CheckedIn, CheckedOut)
        statuses_to_include = []
        if self.chk_registered.isChecked():
            statuses_to_include.append('Registered')
        if self.chk_checkin.isChecked():
            statuses_to_include.append('CheckedIn')
        if self.chk_checkout.isChecked():
            statuses_to_include.append('CheckedOut')
            
        filtered_logs = [row for row in all_logs if row['status'] in statuses_to_include]
        return filtered_logs

    def trigger_export(self, format_type):
        """Distribute execution based on requested file output format."""
        records = self.fetch_records_for_report()
        if not records:
            QMessageBox.warning(self, "No Data", "No visitor records matched the selected filters.")
            return

        os.makedirs(REPORTS_DIR, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        scope_name = ["Daily", "Weekly", "Monthly", "Custom"][self.scope_combo.currentIndex()]
        
        filename = f"VisitorReport_{scope_name}_{timestamp}"

        try:
            if format_type == "CSV":
                filepath = self.export_to_csv(records, filename)
            elif format_type == "EXCEL":
                filepath = self.export_to_excel(records, filename)
            else: # PDF
                filepath = self.export_to_pdf(records, filename)

            # Log audit event
            AuditLogModel.log_event(self.user_session['id'], "Export Report", 
                                  f"Exported {len(records)} records in {format_type} format: {os.path.basename(filepath)}")

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Export Complete")
            msg.setText(f"Report exported successfully!\n\nSaved to: {filepath}")
            
            open_btn = msg.addButton("Open File", QMessageBox.ActionRole)
            msg.addButton(QMessageBox.Close)
            
            msg.exec_()
            
            if msg.clickedButton() == open_btn:
                os.startfile(filepath)

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to generate report file: {e}")

    def export_to_csv(self, records, filename):
        """Write records to a CSV file."""
        filepath = os.path.join(REPORTS_DIR, f"{filename}.csv")
        headers = [
            "Visitor ID", "Full Name", "Mobile", "Company Name", 
            "Purpose", "Host Employee", "Department", "Check-In Time", 
            "Check-Out Time", "Status"
        ]

        with open(filepath, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in records:
                writer.writerow([
                    r['id'], r['full_name'], r['mobile'], r.get('company_name', 'N/A'),
                    r['purpose'], r.get('employee_name', 'N/A'), r.get('department_name', 'N/A'),
                    r.get('check_in_time', 'N/A'), r.get('check_out_time', 'N/A'), r['status']
                ])
        return filepath

    def export_to_excel(self, records, filename):
        """Generate a styled corporate spreadsheet using openpyxl."""
        filepath = os.path.join(REPORTS_DIR, f"{filename}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitor Log"
        
        # Grid lines visible
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10)
        fill_header = PatternFill(start_color="1a252f", end_color="1a252f", fill_type="solid")
        fill_zebra = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        border_thin = Side(style='thin', color='cbd5e1')
        border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        # Title block
        ws.merge_cells("A1:J1")
        ws["A1"] = f"SMARTVMS SYSTEM REPORT - {self.scope_combo.currentText().upper()}"
        ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="00adb5")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # Subtitle details
        ws.merge_cells("A2:J2")
        ws["A2"] = f"Generated by: {self.user_session['full_name']}  |  Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # Empty row spacer
        ws.row_dimensions[3].height = 10

        # Column Headers
        headers = [
            "Visitor ID", "Full Name", "Mobile", "Company Name", 
            "Purpose of Visit", "Host Employee", "Department", 
            "Check-In Time", "Check-Out Time", "Status"
        ]
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws.row_dimensions[4].height = 25

        # Data rows
        for row_idx, r in enumerate(records, 5):
            row_data = [
                r['id'], r['full_name'], r['mobile'], r.get('company_name', 'N/A'),
                r['purpose'], r.get('employee_name', 'N/A'), r.get('department_name', 'N/A'),
                r.get('check_in_time', 'N/A'), r.get('check_out_time', 'N/A'), r['status']
            ]
            
            is_even = (row_idx % 2 == 0)
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_all
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Apply alternate shading
                if is_even:
                    cell.fill = fill_zebra
                    
                # Center align ID and Status
                if col_idx in [1, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
            ws.row_dimensions[row_idx].height = 20

        # Adjust columns width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        wb.save(filepath)
        return filepath

    def export_to_pdf(self, records, filename):
        """Generate a landscape PDF report log using reportlab."""
        filepath = os.path.join(REPORTS_DIR, f"{filename}.pdf")
        
        # Set landscape pagesize
        doc = SimpleDocTemplate(filepath, pagesize=landscape(letter),
                                rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#1a252f'),
            spaceAfter=4
        )
        
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=9,
            textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=15
        )

        header_cell_style = ParagraphStyle(
            'HeaderCell',
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white,
            alignment=1 # Center
        )

        body_cell_style = ParagraphStyle(
            'BodyCell',
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#2c3e50')
        )

        body_cell_center = ParagraphStyle(
            'BodyCellCenter',
            parent=body_cell_style,
            alignment=1 # Center
        )

        elements = []
        
        # Titles
        elements.append(Paragraph(f"SmartVMS - Visitor Logs Report ({self.scope_combo.currentText()})", title_style))
        gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generated by: {self.user_session['full_name']}  |  Timestamp: {gen_time}  |  Total Count: {len(records)}", subtitle_style))
        
        # Build Table Data
        # Column Headers
        table_data = [[
            Paragraph("<b>Visitor ID</b>", header_cell_style),
            Paragraph("<b>Full Name</b>", header_cell_style),
            Paragraph("<b>Mobile</b>", header_cell_style),
            Paragraph("<b>Company</b>", header_cell_style),
            Paragraph("<b>Purpose</b>", header_cell_style),
            Paragraph("<b>Host</b>", header_cell_style),
            Paragraph("<b>Check-In</b>", header_cell_style),
            Paragraph("<b>Check-Out</b>", header_cell_style),
            Paragraph("<b>Status</b>", header_cell_style),
        ]]

        for r in records:
            table_data.append([
                Paragraph(str(r['id']), body_cell_center),
                Paragraph(r['full_name'], body_cell_style),
                Paragraph(r['mobile'], body_cell_center),
                Paragraph(r.get('company_name') or "N/A", body_cell_style),
                Paragraph(r['purpose'], body_cell_center),
                Paragraph(r.get('employee_name') or "N/A", body_cell_style),
                Paragraph(r.get('check_in_time') or "N/A", body_cell_center),
                Paragraph(r.get('check_out_time') or "N/A", body_cell_center),
                Paragraph(f"<b>{r['status']}</b>", body_cell_center),
            ])

        # Width fits 750 pt printable width (Landscape Letter is 792 wide - 40 margins = 752 printable)
        col_widths = [85, 95, 75, 85, 65, 85, 100, 100, 60]
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a252f')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1a252f')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
        ]))

        elements.append(t)
        doc.build(elements)
        return filepath
