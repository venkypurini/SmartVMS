import os
import sys
import threading
import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
from werkzeug.serving import make_server

# Add parent to path if needed so we can import models
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Redirect Flask logs and errors to a persistent file when running inside the EXE
if getattr(sys, 'frozen', False):
    from config import EXECUTABLE_DIR
    log_file = open(os.path.join(EXECUTABLE_DIR, "flask_server.log"), "a", encoding="utf-8", buffering=1)
    sys.stdout = log_file
    sys.stderr = log_file

from database.db_manager import get_db_connection
from models.employee import EmployeeModel
from models.checkin import VisitModel
from models.visitor import VisitorModel
from utils.email_sender import send_approval_request_to_employee

# Initialize Flask App
# The templates folder is explicitly specified so PyInstaller can bundle it later
template_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'templates')
app = Flask(__name__, template_folder=template_dir)
app.config['PUBLIC_URL'] = None
app.secret_key = 'smartvms_admin_session_secret_key_98765'

class ServerThread(threading.Thread):
    def __init__(self, app, host, port):
        threading.Thread.__init__(self)
        self.server = make_server(host, port, app)
        self.ctx = app.app_context()
        self.ctx.push()

    def run(self):
        print(f"[Kiosk] Starting local web server on port {self.server.server_port}...")
        self.server.serve_forever()

    def shutdown(self):
        self.server.shutdown()

@app.route('/')
def index():
    # Redirect root to admin dashboard
    return redirect(url_for('admin_dashboard'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        # Fetch hosts/departments for the form dropdowns
        depts = EmployeeModel.get_all_departments()
        hosts = EmployeeModel.get_all_employees()
        return render_template('register.html', depts=depts, hosts=hosts)
        
    elif request.method == 'POST':
        # Process form submission
        data = request.form
        visitor_name = data.get('visitor_name')
        mobile = data.get('mobile')
        email = data.get('email')
        company = data.get('company')
        purpose = data.get('purpose')
        host_id = data.get('host_id')
        
        try:
            conn = get_db_connection()
            # 1. Figure out dept_id from host
            emp = conn.execute("SELECT department_id FROM employees WHERE id = ?", (host_id,)).fetchone()
            dept_id = emp['department_id'] if emp else 1
            
            # 2. Check if visitor exists (by mobile or email)
            visitor = conn.execute("SELECT id, visitor_code FROM visitors WHERE mobile = ?", (mobile,)).fetchone()
            visitor_by_email = None
            if email:
                visitor_by_email = conn.execute("SELECT id, visitor_code FROM visitors WHERE email = ?", (email,)).fetchone()

            existing = visitor or visitor_by_email

            if existing:
                visitor_id = existing['id']
                visitor_code = existing['visitor_code']
                # Update name, email, etc.
                conn.execute("""
                    UPDATE visitors 
                    SET full_name = ?, email = ?
                    WHERE id = ?;
                """, (visitor_name, email, visitor_id))
                
                # Cancel any previous visits that are pending or approved but not active/completed
                conn.execute("""
                    UPDATE visits
                    SET status = 'cancelled', approval_status = 'rejected'
                    WHERE visitor_id = ? AND status = 'cancelled';
                """, (visitor_id,))
                
                # If matched by email but not mobile, try updating the mobile as well
                if not visitor and visitor_by_email:
                    try:
                        conn.execute("UPDATE visitors SET mobile = ? WHERE id = ?;", (mobile, visitor_id))
                    except Exception:
                        pass
            else:
                # Generate unique visitor code and insert directly
                visitor_code = VisitorModel.generate_visitor_code()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO visitors (
                        visitor_code, full_name, gender, mobile, email, company, photo_path
                    ) VALUES (?, ?, 'Male', ?, ?, ?, 'id_document.jpg');
                """, (visitor_code, visitor_name, mobile, email, company))
                visitor_id = cursor.lastrowid
                
            # 3. Create the pending visit record (waiting for approval)
            now = datetime.datetime.now()
            entry_date = now.strftime("%Y-%m-%d")
            entry_time = now.strftime("%H:%M:%S")
            
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO visits (
                    visitor_id, employee_id, department_id, purpose,
                    entry_date, entry_time, status, approval_status, host_notified
                ) VALUES (?, ?, ?, ?, ?, ?, 'cancelled', 'pending', 0);
            """, (visitor_id, host_id, dept_id, purpose, entry_date, entry_time))
            visit_id = cursor.lastrowid
            
            conn.commit()
            conn.close()
            
            # 4. Send approval request email to employee (host) in a background thread
            try:
                emp_details = EmployeeModel.get_employee_by_id(host_id)
                if emp_details and emp_details['email']:
                    import threading
                    email_thread = threading.Thread(
                        target=send_approval_request_to_employee,
                        args=(
                            emp_details['email'], emp_details['name'],
                            visitor_name, mobile, company, purpose,
                            visitor_code, entry_date
                        ),
                        daemon=True
                    )
                    email_thread.start()
            except Exception as e:
                print(f"[Email Error] Failed to send approval request: {e}")
            
            return jsonify({"status": "success", "message": "Check-in successful! Please wait for approval.", "visit_id": visit_id})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)})

@app.route('/web_approve/<int:visit_id>')
def web_approve(visit_id):
    try:
        from database.db_manager import get_db_connection
        from models.checkin import VisitModel
        from models.visitor import VisitorModel
        from models.audit_logs import AuditLogModel
        from modules.qr_code import QRPassGenerator
        from utils.email_sender import send_qr_to_visitor
        import datetime

        # 1. Fetch visit details
        visit = VisitModel.get_visit_by_id(visit_id)
        if not visit:
            return render_template('approval_result.html',
                                   status='error',
                                   title='Visit Not Found',
                                   visitor_name='',
                                   message='The request details could not be found in the database. Please contact system administration.')

        visitor_name = visit.get('visitor_name', 'Visitor')
        
        # 2. Check current approval status
        current_status = visit.get('approval_status')
        if current_status == 'approved':
            return render_template('approval_result.html',
                                   status='approved',
                                   title='Already Approved',
                                   visitor_name=visitor_name,
                                   message='This visit request has already been approved. The visitor has been emailed their QR pass.')
        elif current_status == 'rejected':
            return render_template('approval_result.html',
                                   status='rejected',
                                   title='Request Rejected',
                                   visitor_name=visitor_name,
                                   message='This visit request has already been rejected. A new request must be submitted to check in.')

        # 3. Fetch visitor company and email details
        visitor_id = visit['visitor_id']
        conn = get_db_connection()
        v_row = conn.execute("SELECT email, company FROM visitors WHERE id = ?", (visitor_id,)).fetchone()
        conn.close()
        visitor_email = v_row['email'] if v_row else ''
        visitor_company = v_row['company'] if v_row else ''

        # 4. Generate QR Code
        qr_path = QRPassGenerator.generate_pass(
            visit['visitor_code'],
            visitor_name,
            visitor_company,
            visit['employee_name'],
            visit['department_name']
        )

        # 5. Approve visit in DB
        VisitorModel.approve_visit(visit_id, approved_by_user_id=None, qr_code_path=qr_path)

        # Log event
        AuditLogModel.log_event(
            user_id=None,
            action="Approve Visit (Web)",
            details=f"Approved visit for {visitor_name} ({visit['visitor_code']}) via host email link.",
            module="Approvals"
        )

        # 6. Send email to visitor
        if visitor_email:
            email_thread = threading.Thread(
                target=send_qr_to_visitor,
                args=(
                    visitor_email,
                    visitor_name,
                    qr_path,
                    visit['employee_name'],
                    visit['department_name'],
                    datetime.date.today().strftime("%d %B %Y")
                ),
                daemon=True
            )
            email_thread.start()
            msg = f"The visit has been approved successfully. A QR pass was generated and sent to the visitor's email address ({visitor_email})."
        else:
            msg = "The visit has been approved successfully. However, the visitor does not have an email address on file, so the QR pass could not be emailed."

        return render_template('approval_result.html',
                               status='approved',
                               title='Visit Request Approved',
                               visitor_name=visitor_name,
                               message=msg)

    except Exception as e:
        print(f"[Web Approve Error] {e}")
        return render_template('approval_result.html',
                               status='error',
                               title='Error Processing Request',
                               visitor_name='',
                               message=f"An unexpected error occurred while processing approval: {str(e)}")

@app.route('/web_reject/<int:visit_id>')
def web_reject(visit_id):
    try:
        from models.checkin import VisitModel
        from models.visitor import VisitorModel
        from models.audit_logs import AuditLogModel

        # 1. Fetch visit details
        visit = VisitModel.get_visit_by_id(visit_id)
        if not visit:
            return render_template('approval_result.html',
                                   status='error',
                                   title='Visit Not Found',
                                   visitor_name='',
                                   message='The request details could not be found in the database. Please contact system administration.')

        visitor_name = visit.get('visitor_name', 'Visitor')

        # 2. Check current approval status
        current_status = visit.get('approval_status')
        if current_status == 'approved':
            return render_template('approval_result.html',
                                   status='approved',
                                   title='Already Approved',
                                   visitor_name=visitor_name,
                                   message='This visit request has already been approved and cannot be rejected now.')
        elif current_status == 'rejected':
            return render_template('approval_result.html',
                                   status='rejected',
                                   title='Already Rejected',
                                   visitor_name=visitor_name,
                                   message='This visit request has already been rejected.')

        # 3. Reject visit in DB
        VisitorModel.reject_visit(visit_id, approved_by_user_id=None)

        # Log event
        AuditLogModel.log_event(
            user_id=None,
            action="Reject Visit (Web)",
            details=f"Rejected visit request for {visitor_name} ({visit['visitor_code']}) via host email link.",
            module="Approvals"
        )

        return render_template('approval_result.html',
                               status='rejected',
                               title='Visit Request Rejected',
                               visitor_name=visitor_name,
                               message='The check-in request has been rejected. The visitor has been denied access to the premises.')

    except Exception as e:
        print(f"[Web Reject Error] {e}")
        return render_template('approval_result.html',
                               status='error',
                               title='Error Processing Request',
                               visitor_name='',
                               message=f"An unexpected error occurred while processing rejection: {str(e)}")

# ==========================================
# WEB ADMIN PORTAL ENDPOINTS
# ==========================================

def is_admin_logged_in():
    return 'user_id' in session

@app.before_request
def restrict_access():
    # Exempt login, logout, static assets, and visitor/host email action routes
    exempt_endpoints = [
        'admin_login',
        'admin_logout',
        'static',
        'web_approve',
        'web_reject',
        'register'
    ]
    if request.endpoint in exempt_endpoints:
        return
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'GET':
        if is_admin_logged_in():
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html')
        
    elif request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        from models.user import UserModel
        user = UserModel.authenticate(username, password)
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            
            # Log audit event
            from models.audit_logs import AuditLogModel
            AuditLogModel.log_event(
                user_id=user['id'],
                action="Admin Web Login",
                details=f"Operator '{user['username']}' logged in via Admin Web Portal.",
                module="Authentication"
            )
            return redirect(url_for('admin_dashboard'))
        else:
            flash("Invalid username or password.", "error")
            return redirect(url_for('admin_login'))

@app.route('/admin/logout')
def admin_logout():
    if is_admin_logged_in():
        from models.audit_logs import AuditLogModel
        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Admin Web Logout",
            details=f"Operator '{session.get('username')}' logged out from Admin Web Portal.",
            module="Authentication"
        )
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_dashboard():
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
        
    from models.visitor import VisitorModel
    from models.checkin import VisitModel
    from database.db_manager import get_db_connection
    import json
    
    # 1. Fetch pending approvals queue
    pending_visits = VisitorModel.get_pending_visits()
    
    # 2. Fetch active visitors
    active_visits = VisitModel.get_visits_history(limit=100, filters={'status': 'active'})
    
    # 3. Fetch recent history
    history = VisitModel.get_visits_history(limit=150)
    
    # 4. Fetch counters
    counters = VisitModel.get_dashboard_counters()
    
    # 5. Fetch employees and departments
    conn = get_db_connection()
    employees = conn.execute("""
        SELECT e.*, d.name as department_name 
        FROM employees e 
        LEFT JOIN departments d ON e.department_id = d.id 
        ORDER BY e.full_name ASC;
    """).fetchall()
    employees_list = [dict(emp) for emp in employees]
    
    departments = conn.execute("SELECT * FROM departments ORDER BY name ASC;").fetchall()
    departments_list = [dict(dept) for dept in departments]
    
    # 6. Fetch audit logs (last 100 entries)
    logs = conn.execute("""
        SELECT a.*, u.full_name as user_fullname 
        FROM audit_logs a 
        LEFT JOIN users u ON a.user_id = u.id 
        ORDER BY a.id DESC LIMIT 100;
    """).fetchall()
    logs_list = [dict(log) for log in logs]
    
    # 7. Fetch system settings
    settings_rows = conn.execute("SELECT key, value FROM system_settings;").fetchall()
    system_settings = {r['key']: r['value'] for r in settings_rows}
    conn.close()
    
    # 8. Fetch SMTP Settings from JSON
    smtp_settings = {}
    smtp_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "database", "smtp_settings.json"))
    if os.path.exists(smtp_path):
        try:
            with open(smtp_path, 'r', encoding='utf-8') as f:
                smtp_settings = json.load(f)
        except Exception:
            pass
            
    session_user = {
        'full_name': session.get('full_name'),
        'role': session.get('role')
    }
    
    return render_template('admin_dashboard.html',
                           pending_visits=pending_visits,
                           active_visits=active_visits,
                           history=history,
                           counters=counters,
                           employees=employees_list,
                           departments=departments_list,
                           audit_logs=logs_list,
                           system_settings=system_settings,
                           smtp_settings=smtp_settings,
                           session_user=session_user)

@app.route('/admin/approve/<int:visit_id>', methods=['POST'])
def admin_approve(visit_id):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    
    try:
        from database.db_manager import get_db_connection
        from models.checkin import VisitModel
        from models.visitor import VisitorModel
        from models.audit_logs import AuditLogModel
        from modules.qr_code import QRPassGenerator
        from utils.email_sender import send_qr_to_visitor
        import datetime
        import threading

        # 1. Fetch visit details
        visit = VisitModel.get_visit_by_id(visit_id)
        if not visit:
            flash("Visit request not found.", "error")
            return redirect(url_for('admin_dashboard'))

        visitor_name = visit.get('visitor_name', 'Visitor')
        
        # 2. Check current approval status
        current_status = visit.get('approval_status')
        if current_status == 'approved':
            flash(f"Visit request for {visitor_name} is already approved.", "info")
            return redirect(url_for('admin_dashboard') + '#approvals')
        elif current_status == 'rejected':
            flash(f"Visit request for {visitor_name} was already rejected.", "error")
            return redirect(url_for('admin_dashboard') + '#approvals')

        # 3. Fetch visitor company and email details
        visitor_id = visit['visitor_id']
        conn = get_db_connection()
        v_row = conn.execute("SELECT email, company FROM visitors WHERE id = ?", (visitor_id,)).fetchone()
        conn.close()
        visitor_email = v_row['email'] if v_row else ''
        visitor_company = v_row['company'] if v_row else ''

        # 4. Generate QR Code
        qr_path = QRPassGenerator.generate_pass(
            visit['visitor_code'],
            visitor_name,
            visitor_company,
            visit['employee_name'],
            visit['department_name']
        )

        # 5. Approve visit in DB
        VisitorModel.approve_visit(visit_id, approved_by_user_id=session.get('user_id'), qr_code_path=qr_path)

        # Log event
        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Approve Visit (Admin Web)",
            details=f"Approved visit request for {visitor_name} ({visit['visitor_code']}) via Web Admin Portal.",
            module="Approvals"
        )

        # 6. Send email to visitor
        if visitor_email:
            email_thread = threading.Thread(
                target=send_qr_to_visitor,
                args=(
                    visitor_email,
                    visitor_name,
                    qr_path,
                    visit['employee_name'],
                    visit['department_name'],
                    datetime.date.today().strftime("%d %B %Y")
                ),
                daemon=True
            )
            email_thread.start()
            flash(f"Successfully approved visit for {visitor_name}. QR pass emailed to {visitor_email}.", "success")
        else:
            flash(f"Successfully approved visit for {visitor_name}. (No visitor email on file)", "success")

    except Exception as e:
        print(f"[Admin Approve Error] {e}")
        flash(f"Error processing approval: {str(e)}", "error")

    return redirect(url_for('admin_dashboard') + '#approvals')

@app.route('/admin/reject/<int:visit_id>', methods=['POST'])
def admin_reject(visit_id):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
        
    try:
        from models.checkin import VisitModel
        from models.visitor import VisitorModel
        from models.audit_logs import AuditLogModel

        # 1. Fetch visit details
        visit = VisitModel.get_visit_by_id(visit_id)
        if not visit:
            flash("Visit request not found.", "error")
            return redirect(url_for('admin_dashboard'))

        visitor_name = visit.get('visitor_name', 'Visitor')

        # 2. Check current approval status
        current_status = visit.get('approval_status')
        if current_status == 'approved':
            flash(f"Visit request for {visitor_name} is already approved and cannot be rejected.", "error")
            return redirect(url_for('admin_dashboard') + '#approvals')
        elif current_status == 'rejected':
            flash(f"Visit request for {visitor_name} was already rejected.", "info")
            return redirect(url_for('admin_dashboard') + '#approvals')

        # 3. Reject visit in DB
        VisitorModel.reject_visit(visit_id, approved_by_user_id=session.get('user_id'))

        # Log event
        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Reject Visit (Admin Web)",
            details=f"Rejected visit request for {visitor_name} ({visit['visitor_code']}) via Web Admin Portal.",
            module="Approvals"
        )
        flash(f"Rejected visitor request for {visitor_name}.", "success")

    except Exception as e:
        print(f"[Admin Reject Error] {e}")
        flash(f"Error processing rejection: {str(e)}", "error")

    return redirect(url_for('admin_dashboard') + '#approvals')

@app.route('/admin/checkout/<int:visit_id>', methods=['POST'])
def admin_checkout(visit_id):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
        
    try:
        from models.checkin import VisitModel
        from models.audit_logs import AuditLogModel
        
        # Fetch visit details
        visit = VisitModel.get_visit_by_id(visit_id)
        if not visit:
            flash("Visit record not found.", "error")
            return redirect(url_for('admin_dashboard'))
            
        visitor_name = visit.get('visitor_name', 'Visitor')
        
        # Check out
        VisitModel.complete_visit(visit_id)
        
        # Log event
        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Check Out (Admin Web)",
            details=f"Checked out visitor {visitor_name} ({visit.get('visitor_code')}) via Web Admin Portal.",
            module="Check-Out"
        )
        
        flash(f"Checked out visitor {visitor_name} successfully.", "success")
        
    except Exception as e:
        print(f"[Admin Checkout Error] {e}")
        flash(f"Error checking out visitor: {str(e)}", "error")
        
    return redirect(url_for('admin_dashboard') + '#active')

@app.route('/admin/history/delete/<int:visit_id>', methods=['POST'])
def admin_delete_visit(visit_id):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        from models.checkin import VisitModel
        from models.audit_logs import AuditLogModel
        visit = VisitModel.get_visit_by_id(visit_id)
        if visit:
            visitor_name = visit.get('visitor_name', 'Visitor')
            VisitModel.delete_visit(visit_id)
            AuditLogModel.log_event(
                user_id=session.get('user_id'),
                action="Delete Visit (Web)",
                details=f"Deleted visit record for {visitor_name} (visit_id={visit_id}).",
                module="History"
            )
            flash(f"Successfully deleted history record for {visitor_name}.", "success")
        else:
            flash("Visit record not found.", "error")
    except Exception as e:
        flash(f"Error deleting record: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#history')

@app.route('/admin/employee/add', methods=['POST'])
def admin_add_employee():
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        from database.db_manager import get_db_connection
        from models.audit_logs import AuditLogModel
        
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        designation = request.form.get('designation')
        dept_id = request.form.get('department_id')

        if not full_name or not email or not phone or not dept_id:
            flash("Missing required fields for employee registration.", "error")
            return redirect(url_for('admin_dashboard') + '#employees')

        if not phone.isdigit() or len(phone) != 10:
            flash("Phone number must be exactly 10 digits.", "error")
            return redirect(url_for('admin_dashboard') + '#employees')

        conn = get_db_connection()
        # Check uniqueness of mobile number
        dup = conn.execute("SELECT id FROM employees WHERE phone = ?;", (phone,)).fetchone()
        if dup:
            conn.close()
            flash("Mobile number already exists for another employee.", "error")
            return redirect(url_for('admin_dashboard') + '#employees')
        conn.close()

        conn = get_db_connection()
        row = conn.execute("SELECT emp_code FROM employees ORDER BY id DESC LIMIT 1;").fetchone()
        if row:
            code = row['emp_code']
            try: num = int(code.replace("EMP", "")) + 1
            except Exception: num = 1
        else:
            num = 1
        emp_code = f"EMP{num:03d}"

        conn.execute("""
            INSERT INTO employees (emp_code, full_name, email, phone, department_id, designation, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1);
        """, (emp_code, full_name, email, phone, dept_id, designation))
        conn.commit()
        conn.close()

        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Add Employee (Web)",
            details=f"Registered employee {full_name} ({emp_code}).",
            module="Employees"
        )
        flash(f"Successfully registered employee {full_name} ({emp_code}).", "success")
    except Exception as e:
        flash(f"Error adding employee: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#employees')

@app.route('/admin/employee/delete/<int:emp_id>', methods=['POST'])
def admin_delete_employee(emp_id):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        from database.db_manager import get_db_connection
        from models.audit_logs import AuditLogModel
        
        conn = get_db_connection()
        emp = conn.execute("SELECT full_name, emp_code FROM employees WHERE id = ?", (emp_id,)).fetchone()
        if emp:
            name = emp['full_name']
            code = emp['emp_code']
            conn.execute("DELETE FROM employees WHERE id = ?", (emp_id,))
            conn.commit()
            conn.close()

            AuditLogModel.log_event(
                user_id=session.get('user_id'),
                action="Delete Employee (Web)",
                details=f"Deleted employee {name} ({code}).",
                module="Employees"
            )
            flash(f"Deleted employee {name} successfully.", "success")
        else:
            conn.close()
            flash("Employee not found.", "error")
    except Exception as e:
        flash(f"Error deleting employee: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#employees')

@app.route('/admin/department/add', methods=['POST'])
def admin_add_department():
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        from database.db_manager import get_db_connection
        from models.audit_logs import AuditLogModel
        
        name = request.form.get('name', '').strip()
        if not name:
            flash("Department name cannot be empty.", "error")
            return redirect(url_for('admin_dashboard') + '#employees')

        conn = get_db_connection()
        exists = conn.execute("SELECT id FROM departments WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
        if exists:
            conn.close()
            flash(f"Department '{name}' already exists.", "error")
            return redirect(url_for('admin_dashboard') + '#employees')

        conn.execute("INSERT INTO departments (name, is_active) VALUES (?, 1);", (name,))
        conn.commit()
        conn.close()

        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Add Department (Web)",
            details=f"Created department '{name}'.",
            module="Employees"
        )
        flash(f"Successfully created department '{name}'.", "success")
    except Exception as e:
        flash(f"Error creating department: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#employees')

@app.route('/admin/settings/update', methods=['POST'])
def admin_update_settings():
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        from database.db_manager import get_db_connection
        from models.audit_logs import AuditLogModel
        import json
        
        # 1. Update SMTP Settings JSON
        smtp_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "database", "smtp_settings.json"))
        smtp_data = {
            "enabled": True if request.form.get('smtp_enabled') == 'on' else False,
            "host": request.form.get('smtp_host', 'smtp.gmail.com'),
            "port": int(request.form.get('smtp_port', 465)),
            "username": request.form.get('smtp_username', ''),
            "password": request.form.get('smtp_password', ''),
            "sender_name": request.form.get('smtp_sender_name', 'SmartVMS')
        }
        with open(smtp_path, 'w', encoding='utf-8') as f:
            json.dump(smtp_data, f, indent=2)

        # 2. Update System Settings in DB
        conn = get_db_connection()
        settings_keys = ['app_name', 'allow_unknown_visitors', 'save_raw_photos', 'auto_checkout_minutes']
        for key in settings_keys:
            val = request.form.get(key)
            if val is not None:
                conn.execute("INSERT OR REPLACE INTO system_settings (key, value) VALUES (?, ?);", (key, val))
        conn.commit()
        conn.close()

        AuditLogModel.log_event(
            user_id=session.get('user_id'),
            action="Update Settings (Web)",
            details="Updated system and SMTP configuration via Web Admin Portal.",
            module="Settings"
        )
        flash("Settings updated successfully.", "success")
    except Exception as e:
        flash(f"Error updating settings: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#settings')

@app.route('/admin/export/<format_type>')
def admin_export_report(format_type):
    if not is_admin_logged_in():
        return redirect(url_for('admin_login'))
    try:
        # Get query parameters
        scope = request.args.get('scope', 'daily')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        status_filter = request.args.getlist('status') # list of checked statuses

        # If empty status list, default to all
        if not status_filter:
            status_filter = ['Registered', 'CheckedIn', 'CheckedOut']

        import datetime
        today = datetime.date.today()
        filters = {}
        if scope == 'daily':
            filters['date_from'] = today.strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        elif scope == 'weekly':
            filters['date_from'] = (today - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        elif scope == 'monthly':
            filters['date_from'] = (today - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
            filters['date_to'] = today.strftime("%Y-%m-%d")
        elif scope == 'custom':
            filters['date_from'] = date_from or today.strftime("%Y-%m-%d")
            filters['date_to'] = date_to or today.strftime("%Y-%m-%d")

        from models.visitor import VisitorModel
        # Pull records
        all_logs = VisitorModel.get_visitors_paginated(limit=5000, offset=0, filters=filters)
        
        # Filter status in python to match exact UI selections
        filtered_status_map = {
            'Registered': 'Registered',
            'CheckedIn': 'CheckedIn',
            'CheckedOut': 'CheckedOut'
        }
        records = [row for row in all_logs if row['status'] in status_filter]

        if not records:
            flash("No visitor records matched the selected filters for export.", "error")
            return redirect(url_for('admin_dashboard') + '#reports')

        import tempfile
        import os
        from flask import send_file

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"VisitorReport_{scope.capitalize()}_{timestamp}"

        # Setup temp path
        temp_dir = tempfile.gettempdir()

        if format_type.upper() == 'CSV':
            import csv
            filepath = os.path.join(temp_dir, f"{filename}.csv")
            headers = [
                "Visitor ID", "Full Name", "Mobile", "Company Name", 
                "Purpose", "Host Employee", "Department", "Check-In Time", 
                "Check-Out Time", "Status"
            ]
            with open(filepath, mode="w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for r in records:
                    writer.writerow([
                        r['id'], r['full_name'], r['mobile'], r.get('company_name', 'N/A'),
                        r['purpose'], r.get('employee_name', 'N/A'), r.get('department_name', 'N/A'),
                        r.get('check_in_time', 'N/A'), r.get('check_out_time', 'N/A'), r['status']
                    ])
            
            # Log audit trail
            from models.audit_logs import AuditLogModel
            AuditLogModel.log_event(
                user_id=session.get('user_id'),
                action="Export Web Report (CSV)",
                details=f"Exported {len(records)} records in CSV format.",
                module="Reports"
            )
            return send_file(filepath, as_attachment=True, download_name=f"{filename}.csv", mimetype='text/csv')

        elif format_type.upper() == 'EXCEL':
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            filepath = os.path.join(temp_dir, f"{filename}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Visitor Log"
            ws.views.sheetView[0].showGridLines = True

            # Styles
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10)
            fill_header = PatternFill(start_color="1a252f", end_color="1a252f", fill_type="solid")
            fill_zebra = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            border_thin = Side(style='thin', color='cbd5e1')
            border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

            # Title block
            ws.merge_cells("A1:J1")
            ws["A1"] = f"SMARTVMS SYSTEM REPORT - {scope.upper()} REPORT"
            ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="00adb5")
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 30

            # Subtitle
            ws.merge_cells("A2:J2")
            user_fullname = session.get('full_name', 'Admin')
            ws["A2"] = f"Generated by: {user_fullname}  |  Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
            ws["A2"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 10

            headers = [
                "Visitor ID", "Full Name", "Mobile", "Company Name", 
                "Purpose of Visit", "Host Employee", "Department", 
                "Check-In Time", "Check-Out Time", "Status"
            ]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_all
            ws.row_dimensions[4].height = 25

            for row_idx, r in enumerate(records, 5):
                row_data = [
                    r['id'], r['full_name'], r['mobile'], r.get('company_name', 'N/A'),
                    r['purpose'], r.get('employee_name', 'N/A'), r.get('department_name', 'N/A'),
                    r.get('check_in_time', 'N/A'), r.get('check_out_time', 'N/A'), r['status']
                ]
                is_even = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_body
                    cell.border = border_all
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if is_even:
                        cell.fill = fill_zebra
                    if col_idx in [1, 8, 9, 10]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[row_idx].height = 20

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

            wb.save(filepath)
            
            # Log audit
            from models.audit_logs import AuditLogModel
            AuditLogModel.log_event(
                user_id=session.get('user_id'),
                action="Export Web Report (Excel)",
                details=f"Exported {len(records)} records in Excel format.",
                module="Reports"
            )
            return send_file(filepath, as_attachment=True, download_name=f"{filename}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        elif format_type.upper() == 'PDF':
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            filepath = os.path.join(temp_dir, f"{filename}.pdf")
            doc = SimpleDocTemplate(filepath, pagesize=landscape(letter),
                                    rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=18,
                textColor=colors.HexColor('#1a252f'),
                spaceAfter=4
            )
            subtitle_style = ParagraphStyle(
                'ReportSubtitle',
                parent=styles['Normal'],
                fontName='Helvetica-Oblique',
                fontSize=9,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=15
            )
            header_cell_style = ParagraphStyle(
                'HeaderCell',
                fontName='Helvetica-Bold',
                fontSize=9,
                textColor=colors.white,
                alignment=1
            )
            body_cell_style = ParagraphStyle(
                'BodyCell',
                fontName='Helvetica',
                fontSize=8,
                textColor=colors.HexColor('#2c3e50')
            )
            body_cell_center = ParagraphStyle(
                'BodyCellCenter',
                parent=body_cell_style,
                alignment=1
            )

            elements = []
            elements.append(Paragraph(f"SmartVMS - Visitor Logs Report ({scope.capitalize()} scope)", title_style))
            user_fullname = session.get('full_name', 'Admin')
            gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            elements.append(Paragraph(f"Generated by: {user_fullname}  |  Timestamp: {gen_time}  |  Total Count: {len(records)}", subtitle_style))

            table_data = [[
                Paragraph("<b>Visitor ID</b>", header_cell_style),
                Paragraph("<b>Full Name</b>", header_cell_style),
                Paragraph("<b>Mobile</b>", header_cell_style),
                Paragraph("<b>Company</b>", header_cell_style),
                Paragraph("<b>Purpose</b>", header_cell_style),
                Paragraph("<b>Host</b>", header_cell_style),
                Paragraph("<b>Check-In</b>", header_cell_style),
                Paragraph("<b>Check-Out</b>", header_cell_style),
                Paragraph("<b>Status</b>", header_cell_style),
            ]]

            for r in records:
                table_data.append([
                    Paragraph(str(r['id']), body_cell_center),
                    Paragraph(r['full_name'], body_cell_style),
                    Paragraph(r['mobile'], body_cell_center),
                    Paragraph(r.get('company_name') or "N/A", body_cell_style),
                    Paragraph(r['purpose'], body_cell_center),
                    Paragraph(r.get('employee_name') or "N/A", body_cell_style),
                    Paragraph(r.get('check_in_time') or "N/A", body_cell_center),
                    Paragraph(r.get('check_out_time') or "N/A", body_cell_center),
                    Paragraph(f"<b>{r['status']}</b>", body_cell_center),
                ])

            col_widths = [85, 95, 75, 85, 65, 85, 100, 100, 60]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a252f')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1a252f')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
            ]))
            elements.append(t)
            doc.build(elements)

            # Log audit
            from models.audit_logs import AuditLogModel
            AuditLogModel.log_event(
                user_id=session.get('user_id'),
                action="Export Web Report (PDF)",
                details=f"Exported {len(records)} records in PDF format.",
                module="Reports"
            )
            return send_file(filepath, as_attachment=True, download_name=f"{filename}.pdf", mimetype='application/pdf')

    except Exception as e:
        flash(f"Error exporting report: {str(e)}", "error")
    return redirect(url_for('admin_dashboard') + '#reports')

import socket
import subprocess
import re
import time

def get_local_ip():
    """Retrieve the local IP address of this machine on the network."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

def start_public_tunnel(port=5000):
    """Starts an SSH reverse tunnel to serveo.net to get a public 4G/5G accessible URL."""
    try:
        import os
        from config import EXECUTABLE_DIR
        log_path = os.path.join(EXECUTABLE_DIR, "ssh_tunnel.log")
        err_path = os.path.join(EXECUTABLE_DIR, "tunnel_error.log")
        
        # Remove old error logs
        if os.path.exists(err_path):
            try: os.remove(err_path)
            except Exception: pass
            
        CREATE_NO_WINDOW = 0x08000000 if os.name == 'nt' else 0
        
        process = subprocess.Popen(
            ["ssh", "-o", "StrictHostKeyChecking=no", "-o", "UserKnownHostsFile=/dev/null", "-o", "ServerAliveInterval=15", "-o", "ServerAliveCountMax=3", "-R", f"80:127.0.0.1:{port}", "serveo.net"],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            creationflags=CREATE_NO_WINDOW
        )
        
        # Wait up to 50 lines for the URL to appear in the output
        url = None
        lines_read = []
        for _ in range(50):
            line = process.stdout.readline()
            if not line: break
            lines_read.append(line)
            
            match = re.search(r'(https://[a-zA-Z0-9-]+\.(serveousercontent\.com|serveo\.net))', line)
            if match:
                url = match.group(1)
                break
                
        if url:
            # Crucial: Start a background thread to continually read stdout so the buffer doesn't fill and block SSH
            import threading
            def drain_stdout():
                try:
                    with open(log_path, "a", encoding="utf-8") as log_file:
                        import datetime
                        log_file.write(f"\n--- Tunnel Started at {datetime.datetime.now()} ---\n")
                        log_file.write(f"Assigned URL: {url}\n")
                        for l in lines_read:
                            log_file.write(f"[SSH Init] {l}")
                        while True:
                            line = process.stdout.readline()
                            if not line:
                                log_file.write("--- Tunnel Stream Closed ---\n")
                                break
                            log_file.write(f"[SSH] {line}")
                            log_file.flush()
                except Exception as e:
                    try:
                        with open(err_path, "a", encoding="utf-8") as f:
                            f.write(f"[Drain Error] {e}\n")
                    except Exception:
                        pass
            threading.Thread(target=drain_stdout, daemon=True).start()
            return process, url
        else:
            # Log standard failure details
            with open(err_path, "w", encoding="utf-8") as f:
                f.write("=== Tunnel Startup Failed ===\n")
                f.write(f"Process Poll status: {process.poll()}\n")
                f.write("Output received:\n")
                f.write("".join(lines_read))
            try: process.terminate()
            except Exception: pass
            return None, None
            
    except Exception as e:
        try:
            from config import EXECUTABLE_DIR
            err_path = os.path.join(EXECUTABLE_DIR, "tunnel_error.log")
            with open(err_path, "a", encoding="utf-8") as f:
                f.write(f"[Tunnel Exception] {e}\n")
        except Exception:
            pass
    return None, None

# Start server function to be called from main.py
def start_kiosk_server(port=5000):
    host = '0.0.0.0'
    server_thread = ServerThread(app, host, port)
    server_thread.daemon = True
    server_thread.start()
    
    # Try getting a public URL for 4G/5G
    tunnel_proc, public_url = start_public_tunnel(port)
    if public_url:
        app.config['PUBLIC_URL'] = public_url
        return server_thread, public_url, tunnel_proc
    
    # Fallback to local IP if tunnel fails
    fallback_url = f"http://{get_local_ip()}:{port}"
    app.config['PUBLIC_URL'] = fallback_url
    return server_thread, fallback_url, None
